from typing import List, Dict
from openpyxl import load_workbook


def parse_excel(path: str) -> List[Dict]:
    """Parse Excel file and return text chunks per sheet."""
    wb = load_workbook(path, data_only=True)
    chunks = []
    for idx, sheet in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet]
        cell_texts = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    cell_texts.append(str(cell))
        text = " ".join(cell_texts).strip()
        if text:
            chunks.append({"page": idx, "text": text})
    return chunks
